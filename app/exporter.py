import base64
import os
import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.base import MIMEBase
from email.mime.text import MIMEText
from email import encoders
import openpyxl
from datetime import datetime
import logging

import requests

logger = logging.getLogger(__name__)

def generate_excel(dates, hours_per_day, filename="/tmp/rapport_export.xlsx", description="Desarrollo", posid="N/A", week_number="N/A", multi_client_data=None):
    """
    Generates an Excel file with the registered hours.
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Rapport Horas"

    # 1. First row: Personalized Title
    user_name = os.getenv("NAME") or os.getenv("EMAIL_ADDRESS_SENDER") or os.getenv("USERNAME") or "USUARIO"
    title_text = f"RAPPORT SEMANA {week_number} - {user_name.upper()}"
    
    ws.append([title_text])
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=5)
    
    # Style title row
    title_cell = ws.cell(row=1, column=1)
    title_cell.font = openpyxl.styles.Font(bold=True, color="FFFFFF", size=12)
    title_cell.fill = openpyxl.styles.PatternFill(start_color="6366F1", end_color="6366F1", fill_type="solid")
    title_cell.alignment = openpyxl.styles.Alignment(horizontal="center")

    # 2. Second row: Column Headers
    headers = ["Fecha", "Día", "Horas", "Proyecto", "Descripción"]
    ws.append(headers)

    # Style headers
    for cell in ws[2]:
        cell.font = openpyxl.styles.Font(bold=True)
        cell.fill = openpyxl.styles.PatternFill(start_color="CCE5FF", end_color="CCE5FF", fill_type="solid")

    day_names = ["Lunes", "Martes", "Miércoles", "Jueves", "Viernes", "Sábado", "Domingo"]
    total_hours = 0
    
    for date_str in sorted(dates):
        date_obj = datetime.fromisoformat(date_str)
        day_name = day_names[date_obj.weekday()]
        
        if multi_client_data:
            # Multi-client: multiple rows per day
            for entry in multi_client_data:
                h = entry.get("hours", 0)
                p = entry.get("project", "N/A")
                d = entry.get("description", "")
                ws.append([date_obj.strftime("%Y-%m-%d"), day_name, h, p, d])
                total_hours += h
        else:
            # Standard single registration
            ws.append([date_obj.strftime("%Y-%m-%d"), day_name, hours_per_day, posid, description])
            total_hours += hours_per_day

    ws.append([])
    ws.append(["TOTAL", "", total_hours])
    
    # Adjust column widths slightly for better readability
    dims = {"A": 15, "B": 12, "C": 10, "D": 25, "E": 60}
    for col, value in dims.items():
        ws.column_dimensions[col].width = value

    # Save
    wb.save(filename)
    return filename, total_hours

def send_email(file_path, total_hours, week_number):
    """
    Sends the generated Excel file via email.
    """
    
    recipient = os.getenv("EMAIL_ADDRESS_RECIPIENT", "[EMAIL_ADDRESS]")
    sender = os.getenv("EMAIL_ADDRESS_SENDER", "[EMAIL_ADDRESS]")
    name = os.getenv("NAME", "JOAQUIN CENTURION")
    subject = f"Semana {week_number} - Horas registradas {total_hours} || {name}"
    
    msg = MIMEMultipart()
    msg['From'] = sender
    msg['To'] = recipient
    msg['Subject'] = subject

    body = f"""
Hola,

Se adjunta el registro de horas de la semana {week_number}.
Total de horas registradas: {total_hours}h.

Nombre y Apellido: {name}
    """
    msg.attach(MIMEText(body, 'plain'))

    # Attachment
    with open(file_path, "rb") as attachment:
        part = MIMEBase('application', 'octet-stream')
        part.set_payload(attachment.read())
        encoders.encode_base64(part)
        part.add_header('Content-Disposition', f"attachment; filename= {os.path.basename(file_path)}")
        msg.attach(part)

    smtp_server = os.getenv("SMTP_SERVER", "localhost")
    smtp_port = int(os.getenv("SMTP_PORT", 1025)) # Default for local testing like MailHog or similar
    smtp_user = os.getenv("SMTP_USER")
    smtp_pass = os.getenv("SMTP_PASS")

    try:
        if smtp_server == "localhost":
            # For testing with local dev mail servers
            with smtplib.SMTP(smtp_server, smtp_port) as server:
                server.send_message(msg)
        else:
            # Production SMTP
            with smtplib.SMTP(smtp_server, smtp_port) as server:
                server.starttls()
                if smtp_user and smtp_pass:
                    server.login(smtp_user, smtp_pass)
                server.send_message(msg)
        logger.info(f"Email sent successfully to {recipient}")
        return True
    except Exception as e:
        logger.error(f"Failed to send email: {e}")
        return False


def send_webhook(file_path: str, week_number: str | int, name: str = "", test_email: bool = False) -> bool:
    """
    Sends the generated Excel file to the configured WEBHOOK URL.

    Payload schema:
        {
            "emailSender": "sender@seidor.es",
            "emailRecepiter": "recipient@seidor.com",
            "fileName": "rapport_sem15_20260418_200000.xlsx",
            "fileData": "<base64-encoded file contents>",
            "week": 15,
            "nameRapport": "JOAQUIN CENTURION",
            "testEmail": false
        }
    """
    webhook_url = os.getenv("WEBHOOK", "").strip()
    if not webhook_url:
        logger.warning("WEBHOOK env var is not set — skipping webhook delivery.")
        return False

    filename = os.path.basename(file_path)
    try:
        week_int = int(week_number)
    except (ValueError, TypeError):
        week_int = 0

    try:
        with open(file_path, "rb") as f:
            file_data = base64.b64encode(f.read()).decode("utf-8")

        name_rapport = name.strip() or os.getenv("NAME", "").strip()

        payload = {
            "emailSender": os.getenv("EMAIL_ADDRESS_SENDER", "").strip(),
            "emailRecepiter": os.getenv("EMAIL_ADDRESS_RECIPIENT", "").strip(),
            "fileName": filename,
            "fileData": file_data,
            "week": week_int,
            "nameRapport": name_rapport,
            "testEmail": test_email,
        }

        response = requests.post(
            webhook_url,
            json=payload,
            timeout=30,
            headers={"Content-Type": "application/json"},
        )
        response.raise_for_status()
        logger.info(f"Webhook delivered successfully (HTTP {response.status_code}).")
        return True
    except Exception as exc:
        logger.error(f"Failed to deliver webhook: {exc}")
        return False


from fpdf import FPDF

def generate_pdf(dates, hours_per_day, filename="/tmp/rapport.pdf", description="Desarrollo", posid="N/A", week_number="N/A", multi_client_data=None):
    """Generates a PDF report with the registered hours."""
    pdf = FPDF()
    pdf.add_page()
    pdf.set_font("Arial", "B", 16)
    
    # Title
    name = os.getenv("NAME", "JOAQUIN CENTURION")
    pdf.cell(0, 10, f"Rapport Semanal - Semana {week_number}", ln=True, align="C")
    pdf.set_font("Arial", "", 12)
    pdf.cell(0, 10, f"Empleado: {name}", ln=True, align="C")
    pdf.ln(10)
    
    # Table Header
    pdf.set_fill_color(99, 102, 241) # Accent color
    pdf.set_text_color(255, 255, 255)
    pdf.set_font("Arial", "B", 10)
    pdf.cell(30, 10, "Fecha", 1, 0, "C", True)
    pdf.cell(30, 10, "Día", 1, 0, "C", True)
    pdf.cell(20, 10, "Horas", 1, 0, "C", True)
    pdf.cell(50, 10, "Proyecto", 1, 0, "C", True)
    pdf.cell(60, 10, "Descripción", 1, 1, "C", True)
    
    # Table Content
    pdf.set_text_color(0, 0, 0)
    pdf.set_font("Arial", "", 9)
    day_names = ["Lunes", "Martes", "Miércoles", "Jueves", "Viernes", "Sábado", "Domingo"]
    total_hours = 0
    
    for date_str in sorted(dates):
        date_obj = datetime.fromisoformat(date_str)
        day_name = day_names[date_obj.weekday()]
        
        if multi_client_data:
            for entry in multi_client_data:
                h = entry.get("hours", 0)
                p = entry.get("project", "N/A")
                d = entry.get("description", "")
                pdf.cell(30, 8, date_obj.strftime("%Y-%m-%d"), 1)
                pdf.cell(30, 8, day_name, 1)
                pdf.cell(20, 8, f"{h}h", 1, 0, "C")
                pdf.cell(50, 8, p, 1)
                pdf.cell(60, 8, d, 1)
                total_hours += h
        else:
            pdf.cell(30, 8, date_obj.strftime("%Y-%m-%d"), 1)
            pdf.cell(30, 8, day_name, 1)
            pdf.cell(20, 8, f"{hours_per_day}h", 1, 0, "C")
            pdf.cell(50, 8, posid, 1)
            pdf.cell(60, 8, description, 1)
            total_hours += hours_per_day
    
    pdf.ln(5)
    pdf.set_font("Arial", "B", 11)
    pdf.cell(0, 10, f"TOTAL HORAS SEMANALES: {total_hours}h", 0, 1, "R")
    
    pdf.output(filename)
    return filename
